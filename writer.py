"""
writer.py — Geração do arquivo Excel final com formatação profissional.

Responsável por:
  - Criar uma aba "<BLOCO> Revisado" para cada bloco comparado.
  - Colorir linhas por Status.
  - Congelar cabeçalho e ativar autofiltro.
  - Criar a aba "Resumo" com estatísticas consolidadas.
  - Usar openpyxl para toda a formatação.
"""

from __future__ import annotations

import logging
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from comparator import (
    COL_CAMPOS,
    COL_STATUS,
    ResultadoBloco,
    STATUS_ALTERADO,
    STATUS_EXCLUIDO,
    STATUS_INCLUIDO,
    STATUS_SEM_ALT,
)
from config import (
    CORES_STATUS,
    COR_CABECALHO,
    COR_CABECALHO_FONTE,
    LARGURA_COLUNA_CAMPOS,
    LARGURA_COLUNA_PADRAO,
    LARGURA_COLUNA_STATUS,
    SUFIXO_ABA,
)

logger = logging.getLogger(__name__)

# Limite de caracteres no nome de aba do Excel
_MAX_NOME_ABA = 31

# Fills de preenchimento (criados uma única vez por performance)
_FILLS: Dict[str, PatternFill] = {
    status: PatternFill(
        start_color=cor, end_color=cor, fill_type="solid"
    )
    for status, cor in CORES_STATUS.items()
}

_FILL_CABECALHO = PatternFill(
    start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid"
)
_FONT_CABECALHO = Font(
    bold=True, color=COR_CABECALHO_FONTE, name="Calibri", size=10
)
_FONT_NORMAL = Font(name="Calibri", size=10)
_FONT_BOLD = Font(bold=True, name="Calibri", size=10)

_BORDA_FINA = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_ALINHAMENTO_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALINHAMENTO_ESQ = Alignment(horizontal="left", vertical="center", wrap_text=False)
_ALINHAMENTO_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Função principal de escrita
# ---------------------------------------------------------------------------

def gerar_excel(
    resultados: List[ResultadoBloco],
    caminho_saida: str,
    label_anterior: str,
    label_atual: str,
) -> None:
    """
    Gera o arquivo Excel final com todas as abas revisadas e o Resumo.

    Parâmetros
    ----------
    resultados : List[ResultadoBloco]
        Lista de resultados da comparação por bloco.
    caminho_saida : str
        Caminho completo do arquivo Excel de saída.
    label_anterior : str
        Rótulo do ano/período anterior (ex: "2024").
    label_atual : str
        Rótulo do ano/período atual (ex: "2025").
    """
    wb = Workbook()
    # Remove a aba padrão criada pelo openpyxl
    wb.remove(wb.active)

    # --- Aba Resumo (criada primeiro para aparecer no início) ---
    ws_resumo = wb.create_sheet(title="Resumo")

    # --- Abas de blocos revisados ---
    for resultado in resultados:
        if resultado.df_resultado.empty:
            logger.info(
                "Bloco '%s' sem dados para exibir. Aba não gerada.",
                resultado.nome_bloco,
            )
            continue

        nome_aba = _truncar_nome_aba(
            f"{resultado.nome_bloco}{SUFIXO_ABA}"
        )
        ws = wb.create_sheet(title=nome_aba)
        _preencher_aba_bloco(ws, resultado, label_anterior, label_atual)
        logger.info("Aba '%s' gerada com sucesso.", nome_aba)

    # --- Finaliza o Resumo ---
    _preencher_aba_resumo(ws_resumo, resultados, label_anterior, label_atual)

    # --- Salva ---
    try:
        wb.save(caminho_saida)
        logger.info("Arquivo salvo em: %s", caminho_saida)
    except PermissionError as exc:
        raise PermissionError(
            f"Não foi possível salvar '{caminho_saida}'. "
            "O arquivo pode estar aberto em outro programa."
        ) from exc


# ---------------------------------------------------------------------------
# Aba por bloco
# ---------------------------------------------------------------------------

def _preencher_aba_bloco(
    ws: Worksheet,
    resultado: ResultadoBloco,
    label_anterior: str,
    label_atual: str,
) -> None:
    """Escreve o DataFrame de resultado em uma aba com formatação completa."""

    df = resultado.df_resultado

    # Renomear colunas _ANTERIOR/_ATUAL para _<label> nos cabeçalhos exibidos
    colunas_exibicao = []
    from config import LABEL_ANTERIOR, LABEL_ATUAL
    for col in df.columns:
        col_exib = col.replace(f"_{LABEL_ANTERIOR}", f"_{label_anterior}") \
                      .replace(f"_{LABEL_ATUAL}",    f"_{label_atual}")
        colunas_exibicao.append(col_exib)

    num_cols = len(df.columns)

    # --- Linha de cabeçalho ---
    for col_idx, nome_col in enumerate(colunas_exibicao, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nome_col)
        cell.fill = _FILL_CABECALHO
        cell.font = _FONT_CABECALHO
        cell.border = _BORDA_FINA
        cell.alignment = _ALINHAMENTO_CENTRO

    # --- Linhas de dados ---
    for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        status = str(row_data.get(COL_STATUS, ""))
        fill = _FILLS.get(status, _FILLS[STATUS_SEM_ALT])

        for col_idx, col_name in enumerate(df.columns, start=1):
            valor = row_data[col_name]
            # Converte NaN para string vazia
            if pd.isna(valor) or str(valor).lower() in ("nan", "none", "null"):
                valor = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=str(valor))
            cell.fill = fill
            cell.font = _FONT_NORMAL
            cell.border = _BORDA_FINA

            # Campos Alterados recebe wrap e largura maior
            if col_name == COL_CAMPOS:
                cell.alignment = _ALINHAMENTO_WRAP
            elif col_name == COL_STATUS:
                cell.alignment = _ALINHAMENTO_CENTRO
            else:
                cell.alignment = _ALINHAMENTO_ESQ

    # --- Ajuste de largura das colunas ---
    for col_idx, col_name in enumerate(df.columns, start=1):
        letra = get_column_letter(col_idx)
        if col_name == COL_STATUS:
            ws.column_dimensions[letra].width = LARGURA_COLUNA_STATUS
        elif col_name == COL_CAMPOS:
            ws.column_dimensions[letra].width = LARGURA_COLUNA_CAMPOS
        else:
            ws.column_dimensions[letra].width = LARGURA_COLUNA_PADRAO

    # --- Freeze panes (congela cabeçalho) ---
    ws.freeze_panes = "A2"

    # --- Autofiltro ---
    ws.auto_filter.ref = ws.dimensions

    # --- Altura da linha de cabeçalho ---
    ws.row_dimensions[1].height = 20

    # --- Nota de chave usada ---
    if resultado.chave_automatica:
        ws.sheet_properties.tabColor = "FFC000"  # laranja = atenção


# ---------------------------------------------------------------------------
# Aba Resumo
# ---------------------------------------------------------------------------

_CABECALHOS_RESUMO = [
    "Bloco",
    "Chaves de Identificação",
    "Qtd Incluídos",
    "Qtd Excluídos",
    "Qtd Alterados",
    "Qtd Sem Alteração",
    "Total Linhas",
    "Chave Automática?",
    "Observação / Avisos",
]

_LARGURAS_RESUMO = [12, 40, 16, 16, 16, 18, 14, 18, 60]


def _preencher_aba_resumo(
    ws: Worksheet,
    resultados: List[ResultadoBloco],
    label_anterior: str,
    label_atual: str,
) -> None:
    """Escreve a aba de Resumo com tabela consolidada."""

    # Título
    ws.merge_cells("A1:I1")
    cell_titulo = ws.cell(
        row=1,
        column=1,
        value=f"REVISÃO ECF — {label_anterior} × {label_atual}",
    )
    cell_titulo.fill = _FILL_CABECALHO
    cell_titulo.font = Font(bold=True, color=COR_CABECALHO_FONTE, name="Calibri", size=14)
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Cabeçalho da tabela
    for col_idx, nome in enumerate(_CABECALHOS_RESUMO, start=1):
        cell = ws.cell(row=2, column=col_idx, value=nome)
        cell.fill = _FILL_CABECALHO
        cell.font = _FONT_CABECALHO
        cell.border = _BORDA_FINA
        cell.alignment = _ALINHAMENTO_CENTRO

    # Dados
    for row_idx, resultado in enumerate(resultados, start=3):
        total = (
            resultado.qtd_incluidos
            + resultado.qtd_excluidos
            + resultado.qtd_alterados
            + resultado.qtd_sem_alteracao
        )

        observacao = resultado.aviso
        if resultado.erros:
            sep = " | " if observacao else ""
            observacao += sep + " | ".join(resultado.erros)

        linha = [
            resultado.nome_bloco,
            ", ".join(resultado.chaves_usadas) if resultado.chaves_usadas else "—",
            resultado.qtd_incluidos,
            resultado.qtd_excluidos,
            resultado.qtd_alterados,
            resultado.qtd_sem_alteracao,
            total,
            "SIM ⚠️" if resultado.chave_automatica else "Não",
            observacao,
        ]

        # Cor de fundo da linha por status dominante
        fill_linha = _determinar_fill_resumo(resultado)

        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill = fill_linha
            cell.font = _FONT_BOLD if col_idx == 1 else _FONT_NORMAL
            cell.border = _BORDA_FINA
            cell.alignment = (
                _ALINHAMENTO_CENTRO if col_idx in (1, 3, 4, 5, 6, 7, 8)
                else _ALINHAMENTO_ESQ
            )

    # Ajuste de largura
    for col_idx, largura in enumerate(_LARGURAS_RESUMO, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    # Freeze no cabeçalho da tabela
    ws.freeze_panes = "A3"

    # Autofiltro na tabela
    ultima_col = get_column_letter(len(_CABECALHOS_RESUMO))
    ultima_linha = 2 + len(resultados)
    ws.auto_filter.ref = f"A2:{ultima_col}{ultima_linha}"


def _determinar_fill_resumo(resultado: ResultadoBloco) -> PatternFill:
    """
    Determina a cor de fundo de uma linha do resumo com base no estado geral
    do bloco (bloco novo, removido, com alterações, etc.).
    """
    aviso = resultado.aviso.upper()
    if "NOVO" in aviso:
        return _FILLS[STATUS_INCLUIDO]
    if "REMOVIDO" in aviso:
        return _FILLS[STATUS_EXCLUIDO]
    if resultado.qtd_alterados > 0:
        return _FILLS[STATUS_ALTERADO]
    if resultado.qtd_incluidos > 0 or resultado.qtd_excluidos > 0:
        fill_neutro = PatternFill(
            start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid"
        )
        return fill_neutro
    return _FILLS[STATUS_SEM_ALT]


# ---------------------------------------------------------------------------
# Utilitários
# ---------------------------------------------------------------------------

def _truncar_nome_aba(nome: str) -> str:
    """Trunca o nome da aba para o limite do Excel (31 caracteres)."""
    if len(nome) > _MAX_NOME_ABA:
        logger.warning(
            "Nome de aba '%s' truncado para %d caracteres.", nome, _MAX_NOME_ABA
        )
        return nome[:_MAX_NOME_ABA]
    return nome
